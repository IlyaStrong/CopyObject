import time
from threading import Event

import pyautogui
import pyperclip
from openpyxl import load_workbook

from exceptions import AbortedException


class CopyFromExcel:
    def __init__(self):
        pass

    def load_workbook(self, file_path):
        self.wb = load_workbook(file_path)

    def close_workbook(self):
        self.wb.close()

    def get_sheets(self):
        sheet_names = self.wb.sheetnames
        sheets = []

        for idx, name in enumerate(sheet_names, start=1):
            sheets.append(f"{idx}. {name}")

        return sheets

    def check_index(self, sheet_number: int):
        return 0 <= sheet_number < len(self.wb.sheetnames)

    def read_excel_column(self, sheet_number: int, column_letter: str, size: int):
        sheet_names = self.wb.sheetnames
        sheet = self.wb[sheet_names[sheet_number]]
        count_times = 1
        current_number = 1
        cell_values = []

        while current_number <= size:
            # Получение значения ячейки
            cell_reference = f"{column_letter}{count_times}"
            cell_value = sheet[cell_reference].value

            # Пропуск итерации, если ячейка содержит None
            if cell_value is None:
                count_times += 1
                continue

            cell_values.append(cell_value)
            count_times += 1
            current_number += 1

        return cell_values

    def paste_values(self, coord, cell_values):
        pyautogui.moveTo(coord)
        pyautogui.click()

        for count, value in enumerate(cell_values):
            if self.cancelled.is_set():
                raise AbortedException()

            if count > 0:
                pyautogui.press("down")

            if self.skip_f2 is False:
                pyautogui.press("f2")

            pyperclip.copy(value)
            pyautogui.hotkey("ctrl", "v")
            pyautogui.press("enter")

            time.sleep(0.1)

    def setup(
        self,
        skip_f2: bool = False,
        cancelled: Event = Event(),
    ):
        self.skip_f2 = skip_f2
        self.cancelled = cancelled
