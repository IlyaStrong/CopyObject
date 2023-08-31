from openpyxl import load_workbook
from pynput.mouse import Listener
import keyboard
import threading
import time
import pyautogui
import openpyxl
import pyperclip 
import sys
stop=False

def exit_on_esc():
    """
    Непрерывно проверяет нажатые клавиши
    В случае нажатия ESC переменная выхода меняется на True и происходит выход
    :return:
    """

    while True:
        key=keyboard.read_key()
        if key in ['esc']:
            global stop
            stop=True
            sys.exit()
            
def get_click_coordinates(message):
    print(message)
    click_point = None

    def on_click(x, y, button, pressed):
        nonlocal click_point
        if pressed:
            click_point = (x, y)
            return False  # Останавливаем слушателя после первого клика

    with Listener(on_click=on_click) as listener:
        listener.join()

    return click_point

def paste_values(coord, cell_values):
    pyautogui.moveTo(coord)
    pyautogui.click()  
    for count, value in enumerate(cell_values):
        if stop == True:
            sys.exit()        
        if count > 0:
            pyautogui.press('down')   
        pyautogui.press('f2')
        pyperclip.copy(value)
        keyboard.press_and_release('ctrl+v')
        pyautogui.press('enter')
        time.sleep(0.1)

def read_excel_column():
    file_path = r'C:\Users\Илья\Desktop\python\testу.xlsx'
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    print("Введите номер листа с которого осуществить копирование")
    print("Доступные листы:")
    for idx, name in enumerate(sheet_names, start=1):
        print(f"{idx}. {name}")
    sheet_number = int(input("Введите номер листа в Excel для чтения: ")) - 1

    # Проверка корректности номера листа
    if 0 <= sheet_number < len(sheet_names):
        sheet = wb[sheet_names[sheet_number]]

        # Запрос у пользователя буквы столбца
        column_letter = input("Введите букву столбца для чтения: ").upper()  # Преобразуем в верхний регистр
        size = int(input("Введите количество строк, которые требуется скопировать: "))

        count_times = 1
        current_number = 1
        cell_values = []
        while current_number <= size:
            # Получение значения ячейки
            cell_reference = f'{column_letter}{count_times}'
            cell_value = sheet[cell_reference].value

            # Пропуск итерации, если ячейка содержит None
            if cell_value is None:
                count_times += 1
                continue
        
            cell_values.append(cell_value)
            count_times += 1
            current_number += 1

        return cell_values
    else:
        print("Неверный номер листа.")

def main():

    coord = get_click_coordinates("Сделайте клик на первый объект списка.")
    cell_values = read_excel_column()   
    paste_values(coord, cell_values)

if __name__ == "__main__":
    exit_thread = threading.Thread(target=exit_on_esc)
    exit_thread.start()
    main_thread = threading.Thread(target=main)
    main_thread.start()
